"""
This module contains function related to generation of excel
file with calendar inside it.
"""

import calendar
import os

import click
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelCalendar():
    """Class representing excel file with the calendar."""

    # pylint: disable=too-few-public-methods

    def __init__(self, date, config, file_path=None):
        self.workbook = Workbook()
        self.date = date
        self.config = config

        if file_path is not None:
            self.file_name = os.path.basename(file_path)
            self.file_path = file_path
        else:
            cwd = os.getcwd()
            date_string = date.strftime("%Y%m")
            self.file_name = "calendar_" + date_string + ".xlsx"
            self.file_path = os.path.join(cwd, self.file_name)

    def _apply_style_header(self, cell, date):
        """Apply header style"""
        weekday = self.config["header"]["weekday"]
        saturday = self.config["header"]["saturday"]
        sunday = self.config["header"]["sunday"]

        alignment = None
        border = None
        font = None
        fill = None

        alignment = Alignment(wrap_text=False,
                              horizontal='center',
                              vertical='center')

        border = Border(top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000'),
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'))

        if date.weekday() == 5:
            font = Font(name=saturday["font"],
                        size=saturday["size"],
                        bold=saturday["bold"],
                        italic=saturday["italic"],
                        color=saturday["color"])
            fill = PatternFill(patternType='solid',
                               fill_type='solid',
                               fgColor=Color(saturday["background"]))
        elif date.weekday() == 6:
            font = Font(name=sunday["font"],
                        size=sunday["size"],
                        bold=sunday["bold"],
                        italic=sunday["italic"],
                        color=sunday["color"])
            fill = PatternFill(patternType='solid',
                               fill_type='solid',
                               fgColor=Color(sunday["background"]))
        else:
            font = Font(name=weekday["font"],
                        size=weekday["size"],
                        bold=weekday["bold"],
                        italic=weekday["italic"],
                        color=weekday["color"])
            fill = PatternFill(patternType='solid',
                               fill_type='solid',
                               fgColor=Color(weekday["background"]))

        cell.alignment = alignment
        cell.border = border
        cell.font = font
        cell.fill = fill

    def _apply_style_date(self, cell, date):
        """Apply date style"""
        weekday = self.config["dates"]["weekday"]
        saturday = self.config["dates"]["saturday"]
        sunday = self.config["dates"]["sunday"]
        other = self.config["dates"]["other"]

        alignment = None
        border = None
        font = None
        fill = None

        alignment = Alignment(wrap_text=True,
                              horizontal='left',
                              vertical='top')

        border = Border(top=Side(style='thin', color='000000'),
                        bottom=Side(style='thin', color='000000'),
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'))

        # Check if the date is within current month
        if date.month != self.date.month:
            font = Font(name=other["font"],
                        size=other["size"],
                        bold=other["bold"],
                        italic=other["italic"],
                        color=other["color"])
            fill = PatternFill(patternType='solid',
                               fill_type='solid',
                               fgColor=Color(other["background"]))
        else:
            font = Font(name=weekday["font"],
                        size=weekday["size"],
                        bold=weekday["bold"],
                        italic=weekday["italic"],
                        color=weekday["color"])

        # Set special background colors for weeknds
        if fill is None:
            if date.weekday() == 5:
                fill = PatternFill(patternType='solid',
                                   fill_type='solid',
                                   fgColor=Color(saturday["background"]))

            elif date.weekday() == 6:
                fill = PatternFill(patternType='solid',
                                   fill_type='solid',
                                   fgColor=Color(sunday["background"]))
            else:
                fill = PatternFill(patternType='solid',
                                   fill_type='solid',
                                   fgColor=Color(weekday["background"]))

        cell.alignment = alignment
        cell.border = border
        cell.font = font
        cell.fill = fill

    def paint_calendar(self):
        """Paint the calender into excel worksheet"""
        if os.path.exists(self.file_path):
            click.echo("Skipping {} as it already exists.".format(
                self.file_name))
            return

        cal = calendar.Calendar(firstweekday=6)
        cal_list = cal.monthdatescalendar(self.date.year, self.date.month)

        worksheet = self.workbook.active
        worksheet.title = self.date.strftime("%Y%m")

        worksheet.row_dimensions[1].height = 22.50
        for column_index, date in enumerate(cal_list[0], 1):

            # Set calendar's column width
            column_letter = get_column_letter(column_index)
            worksheet.column_dimensions[column_letter].width = 10.38

            weekday_index = date.weekday()
            weekday_list = ["月", "火", "水", "木", "金", "土", "日"]
            current_cell = worksheet.cell(column=column_index,
                                          row=1,
                                          value=weekday_list[weekday_index])
            self._apply_style_header(current_cell, date)

        for row_index, row in enumerate(cal_list, 2):
            worksheet.row_dimensions[row_index].height = 54.75
            for column_index, date in enumerate(row, 1):
                if date.month != self.date.month:
                    if os.name == 'nt':
                        val = date.strftime("%#m/%#d") + "\n"
                    else:
                        val = date.strftime("%-m/%-d") + "\n"
                else:
                    if os.name == 'nt':
                        val = date.strftime("%#d") + "\n"
                    else:
                        val = date.strftime("%-d") + "\n"
                current_cell = worksheet.cell(column=column_index,
                                              row=row_index,
                                              value=val)
                self._apply_style_date(current_cell, date)

        self.workbook.save(filename=self.file_path)
